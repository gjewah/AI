# -*- coding: utf-8 -*-
import io
from odoo.tests.common import HttpCase, tagged
from openpyxl import load_workbook
@tagged('-at_install','post_install')
class TestESGExportFallback(HttpCase):
    def test_export_uses_dpp_default_when_esg_empty(self):
        Ind=self.env['vidir.industry']; Prof=self.env['vidir.industry.profile']; Fac=self.env['vidir.industry.factor.set']
        ind=Ind.create({'name':'Core','code':'core'})
        prof=Prof.create({'name':'Core-NO-2026','industry_id':ind.id,'country':'NO','year':2026,'company_id':self.env.company.id})
        Fac.create({'profile_id':prof.id,'code':'grid_NO','unit':'kg/kWh','value':0.20})
        Fac.create({'profile_id':prof.id,'code':'marine_fuel_ton_hfo','unit':'kg/ton','value':3100})
        dpp=self.env['vidir.dpp.passport'].create({'name':'Vessel','product_tmpl_id': self.env['product.template'].create({'name':'Ship'}).id,'country_of_origin':'NO','industry_id':ind.id,'industry_profile_id':prof.id,'default_marine_fuel_type':'hfo','default_marine_fuel_ton':1.5})
        r=self.url_open('/dpp/%s/export/xlsx' % dpp.uuid)
        self.assertEqual(r.status_code, 200)
        wb=load_workbook(io.BytesIO(r.content)); ws=wb.active
        found=False
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[2]=='marine_fuel_ton_hfo':
                self.assertEqual(row[3],'ton'); self.assertAlmostEqual(float(row[4]), 1.5, places=6); self.assertAlmostEqual(float(row[5]), 4650.0, places=6)
                found=True; break
        self.assertTrue(found)
