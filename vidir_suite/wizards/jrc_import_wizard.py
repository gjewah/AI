
# -*- coding: utf-8 -*-
import io, json, base64
from odoo import api, fields, models, _
from odoo.exceptions import UserError
try:
    import openpyxl
except Exception:
    openpyxl=None
class VidirJRCImportWizard(models.TransientModel):
    _name='vidir.jrc.import.wizard'; _description='JRC/NEEFE Excel-import'
    file_data=fields.Binary(string='Excel-fil (XLSX)', required=True)
    file_name=fields.Char(); sheet_name=fields.Char(string='Ark (valgfritt)')
    country_col=fields.Char(default='A'); factor_col=fields.Char(default='B')
    def action_import(self):
        if not openpyxl: raise UserError(_('openpyxl mangler; legg i miljøet'))
        if not self.file_data: raise UserError(_('Mangler fil'))
        content=io.BytesIO(base64.b64decode(self.file_data)); wb=openpyxl.load_workbook(content,data_only=True)
        ws=wb[self.sheet_name] if self.sheet_name and self.sheet_name in wb.sheetnames else wb.active
        def col_idx(c):
            s=0
            for ch in c.strip().upper():
                if 'A' <= ch <= 'Z': s=s*26+(ord(ch)-64)
            return s
        ci=col_idx(self.country_col); fi=col_idx(self.factor_col)
        data={}
        for r in ws.iter_rows(min_row=2):
            cc=r[ci-1].value; fv=r[fi-1].value
            if not cc: continue
            try:
                cc=str(cc).strip().upper(); val=float(str(fv).replace(',','.'))
                if val>0: data[cc]=val
            except: pass
        if not data: raise UserError(_('Ingen gyldige rader'))
        self.env['ir.config_parameter'].sudo().set_param('vidir_co2.grid_factors_json', json.dumps(data))
        return {'type':'ir.actions.client','tag':'display_notification','params':{'title':_('Import OK'),'message':_('Lest %s land')%len(data),'sticky':False}}
