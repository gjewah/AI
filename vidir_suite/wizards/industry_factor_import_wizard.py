
# -*- coding: utf-8 -*-
import base64, io
from odoo import api, fields, models, _
from odoo.exceptions import UserError
try:
    import openpyxl
except Exception:
    openpyxl=None

class VidirIndustryFactorImportWizard(models.TransientModel):
    _name = 'vidir.industry.factor.import.wizard'
    _description = 'Importer faktorer til profil (XLSX)'

    file_data = fields.Binary(string='Excel (XLSX)', required=True)
    file_name = fields.Char()
    industry_id = fields.Many2one('vidir.industry', required=True)
    profile_id = fields.Many2one('vidir.industry.profile', string='Målprofil', required=True, domain="[('industry_id','=',industry_id),('company_id','=',company_id)]")

    def action_import(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('openpyxl mangler – installer i miljøet'))
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(self.file_data)), data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            code_idx = headers.index('code'); unit_idx = headers.index('unit'); value_idx = headers.index('value')
        except ValueError:
            raise UserError(_('Første rad må inneholde kolonnene: code, unit, value'))
        Factor = self.env['vidir.industry.factor.set'].sudo()
        count=0
        for row in ws.iter_rows(min_row=2):
            code = (row[code_idx].value or '').strip()
            unit = (row[unit_idx].value or '')
            value = row[value_idx].value or 0.0
            if not code:
                continue
            rec = Factor.search([('profile_id','=',self.profile_id.id),('code','=',code)], limit=1)
            vals={'profile_id': self.profile_id.id, 'code': code, 'unit': unit, 'value': float(value)}
            if rec:
                rec.write(vals)
            else:
                Factor.create(vals)
            count += 1
        return {'type':'ir.actions.client','tag':'display_notification','params':{'title':_('Importert'),'message':_('Oppdaterte %s rader')%count,'sticky':False}}
